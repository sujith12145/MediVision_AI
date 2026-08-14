"""
Finance router — POST /api/finance, GET /api/finance, GET /api/finance/overview/{month}
"""

import re
import io
import json
from typing import Any
from datetime import date, datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db  # type: ignore
from app.dependencies import get_current_user, SupabaseUser  # type: ignore
from app.models.monthly_finance import MonthlyFinance  # type: ignore
from app.models.medicine import Medicine  # type: ignore
from app.models.sale import Sale  # type: ignore
from app.models.sale_item import SaleItem  # type: ignore
from app.models.extraction_record import ExtractionRecord  # type: ignore
from app.models.audit_log import AuditLog  # type: ignore


router = APIRouter(prefix="/finance", tags=["finance"])

# ── Schemas ──────────────────────────────────────────────────────────────────

class MonthlyFinanceResponse(BaseModel):
    id: int
    month: str
    rent: float
    electricity_and_bills: float
    staff_salaries: float
    other_expenses: float
    other_revenue: float
    computed_revenue: float
    total_revenue: float
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class MonthlyFinanceOverviewResponse(BaseModel):
    month: str
    rent: float
    electricity_and_bills: float
    staff_salaries: float
    other_expenses: float
    other_revenue: float
    computed_revenue: float
    total_revenue: float
    total_costs: float
    current_inventory_investment: float
    net_profit: float
    return_on_investment: float # ROI as percentage (e.g. 12.5)
    estimated_margin: float



class MonthlyFinanceSaveRequest(BaseModel):
    # Server-side validation: month must match YYYY-MM and numbers must be non-negative
    month: str = Field(..., description="Month in YYYY-MM format")
    rent: float = Field(default=0.0, ge=0.0, description="Rent expense")
    electricity_and_bills: float = Field(default=0.0, ge=0.0, description="Electricity & other bills")
    staff_salaries: float = Field(default=0.0, ge=0.0, description="Staff salaries")
    other_expenses: float = Field(default=0.0, ge=0.0, description="Other fixed expenses")
    other_revenue: float = Field(default=0.0, ge=0.0, description="Other sales/revenue manually entered")


# ── Endpoints ────────────────────────────────────────────────────────────────

def get_computed_sales_for_month(db: Session, month: str) -> float:
    """
    Computes total sales for a given month format YYYY-MM by querying the sales table.
    """
    try:
        year, month_num = map(int, month.split("-"))
        import calendar
        last_day = calendar.monthrange(year, month_num)[1]
        
        # Start and end of the month
        start_dt = datetime(year, month_num, 1, 0, 0, 0)
        end_dt = datetime(year, month_num, last_day, 23, 59, 59, 999999)
        
        sales_val = db.query(func.sum(Sale.total_amount)).filter(
            Sale.sold_at >= start_dt,
            Sale.sold_at <= end_dt
        ).scalar()
        
        return float(sales_val) if sales_val is not None else 0.0
    except Exception:
        return 0.0


@router.get(
    "",
    response_model=list[MonthlyFinanceResponse],
    summary="Get all monthly finance records",
    description="Returns a list of all saved monthly finance records, sorted by month descending.",
)
def get_finance_records(
    db: Session = Depends(get_db),
    current_user: SupabaseUser = Depends(get_current_user),
):
    if current_user.role not in ("admin", "pharmacist"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Staff cannot access Monthly Business Overview."
        )
    records = (
        db.query(MonthlyFinance)
        .order_by(MonthlyFinance.month.desc())
        .all()
    )
    
    results = []
    for record in records:
        computed_sales = get_computed_sales_for_month(db, record.month)
        other_rev = float(record.other_revenue)
        total_rev = computed_sales + other_rev
        
        results.append({
            "id": record.id,
            "month": record.month,
            "rent": float(record.rent),
            "electricity_and_bills": float(record.electricity_and_bills),
            "staff_salaries": float(record.staff_salaries),
            "other_expenses": float(record.other_expenses),
            "other_revenue": other_rev,
            "computed_revenue": computed_sales,
            "total_revenue": total_rev,
            "created_at": record.created_at,
            "updated_at": record.updated_at,
        })
    return results


@router.get(
    "/overview/{month}",
    response_model=MonthlyFinanceOverviewResponse,
    summary="Get financial metrics overview for a specific month",
    description="Returns detailed cost, revenue, live inventory investment, net profit, and ROI % for a given month.",
)
def get_finance_overview(
    month: str,
    db: Session = Depends(get_db),
    current_user: SupabaseUser = Depends(get_current_user),
) -> MonthlyFinanceOverviewResponse:
    if current_user.role not in ("admin", "pharmacist"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Staff cannot access Monthly Business Overview."
        )
    # Validate month format (YYYY-MM)
    if not re.match(r"^\d{4}-\d{2}$", month):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Month must be in YYYY-MM format (e.g., 2026-07)"
        )

    # Fetch monthly finance entry
    record = (
        db.query(MonthlyFinance)
        .filter(MonthlyFinance.month == month)
        .first()
    )

    rent = float(record.rent) if record else 0.0
    bills = float(record.electricity_and_bills) if record else 0.0
    salaries = float(record.staff_salaries) if record else 0.0
    other = float(record.other_expenses) if record else 0.0
    other_revenue = float(record.other_revenue) if record else 0.0

    # Live query calculated sales for this month
    computed_revenue = get_computed_sales_for_month(db, month)
    total_revenue = computed_revenue + other_revenue

    # 1. Total costs = rent + bills + salaries + other_expenses
    total_costs = rent + bills + salaries + other

    # 2. Current inventory investment = sum of (quantity * purchase_price) across in-stock medicines
    # Using secure parameterized sum
    inventory_val = db.query(func.sum(Medicine.quantity * Medicine.purchase_price)).filter(Medicine.quantity > 0).scalar()
    current_inventory_investment = float(inventory_val) if inventory_val is not None else 0.0

    # 3. Net profit/loss = total_revenue - total_costs
    net_profit = total_revenue - total_costs

    # 4. Return on investment = net profit / current inventory investment
    if current_inventory_investment > 0.0:
        return_on_investment = (net_profit / current_inventory_investment) * 100.0
    else:
        return_on_investment = 0.0

    # 5. Estimated Margin = sum of (sale_item.sale_price - medicine.purchase_price) * sale_item.quantity_sold
    year_str, month_str = month.split("-")
    year = int(year_str)
    month_num = int(month_str)
    start_dt = datetime(year, month_num, 1, 0, 0, 0)
    if month_num == 12:
        end_dt = datetime(year + 1, 1, 1, 0, 0, 0)
    else:
        end_dt = datetime(year, month_num + 1, 1, 0, 0, 0)

    margin_val = (
        db.query(func.sum((SaleItem.sale_price - Medicine.purchase_price) * SaleItem.quantity_sold))
        .select_from(SaleItem)
        .join(Sale, SaleItem.sale_id == Sale.id)
        .join(Medicine, SaleItem.medicine_id == Medicine.id)
        .filter(
            Sale.sold_at >= start_dt,
            Sale.sold_at < end_dt
        )
        .scalar()
    )
    estimated_margin = float(margin_val) if margin_val is not None else 0.0

    return MonthlyFinanceOverviewResponse(
        month=month,
        rent=rent,
        electricity_and_bills=bills,
        staff_salaries=salaries,
        other_expenses=other,
        other_revenue=other_revenue,
        computed_revenue=computed_revenue,
        total_revenue=total_revenue,
        total_costs=total_costs,
        current_inventory_investment=current_inventory_investment,
        net_profit=net_profit,
        return_on_investment=return_on_investment,
        estimated_margin=estimated_margin
    )


@router.post(
    "",
    response_model=MonthlyFinanceResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Save or update a monthly finance record",
    description="Saves a financial entry for a month. Creates a new record or updates the existing one.",
)
def save_finance_record(
    request_data: MonthlyFinanceSaveRequest,
    db: Session = Depends(get_db),
    current_user: SupabaseUser = Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can edit Monthly Business Overview entries."
        )
    # Validate month format
    if not re.match(r"^\d{4}-\d{2}$", request_data.month):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Month must be in YYYY-MM format (e.g., 2026-07)"
        )

    # Fetch existing
    record = (
        db.query(MonthlyFinance)
        .filter(MonthlyFinance.month == request_data.month)
        .first()
    )

    if record:
        # Update existing
        record.rent = request_data.rent
        record.electricity_and_bills = request_data.electricity_and_bills
        record.staff_salaries = request_data.staff_salaries
        record.other_expenses = request_data.other_expenses
        record.other_revenue = request_data.other_revenue
    else:
        # Create new
        record = MonthlyFinance(
            month=request_data.month,
            rent=request_data.rent,
            electricity_and_bills=request_data.electricity_and_bills,
            staff_salaries=request_data.staff_salaries,
            other_expenses=request_data.other_expenses,
            other_revenue=request_data.other_revenue
        )
        db.add(record)

    db.commit()
    db.refresh(record)

    computed_sales = get_computed_sales_for_month(db, record.month)
    other_rev = float(record.other_revenue)
    total_rev = computed_sales + other_rev

    return {
        "id": record.id,
        "month": record.month,
        "rent": float(record.rent),
        "electricity_and_bills": float(record.electricity_and_bills),
        "staff_salaries": float(record.staff_salaries),
        "other_expenses": float(record.other_expenses),
        "other_revenue": other_rev,
        "computed_revenue": computed_sales,
        "total_revenue": total_rev,
        "created_at": record.created_at,
        "updated_at": record.updated_at,
    }


# ── PDF / Excel Helpers for Reports ──────────────────────────────────────────

def generate_gst_excel_report(month: str, data_rows: list, total_taxable: float, total_tax: float, total_grand: float) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws: Any = wb.active
    ws.title = "GST Report"
    
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Segoe UI", size=14, bold=True, color="1F497D")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    disclaimer_font = Font(name="Segoe UI", size=9, italic=True, color="C00000")
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    summary_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # Title Block
    ws.cell(row=1, column=1, value="MediVision AI - GST Sales Report").font = title_font
    ws.cell(row=2, column=1, value=f"Report Period: {month}").font = bold_font
    ws.cell(row=3, column=1, value="This report is for reference only. Verify all figures and file directly through the official GST portal or your accountant.").font = disclaimer_font
    
    # Headers
    headers = [
        "Date", "Medicine Name", "HSN Code", "Quantity", 
        "Sale Price", "Taxable Value", "GST Rate (%)", "Tax Amount", "Total Amount"
    ]
    
    header_row = 5
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Data Rows
    current_row = header_row + 1
    for r in data_rows:
        ws.cell(row=current_row, column=1, value=r["date"]).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=2, value=r["name"]).alignment = Alignment(horizontal="left")
        
        hsn_cell = ws.cell(row=current_row, column=3, value=r["hsn"])
        hsn_cell.alignment = Alignment(horizontal="center")
        hsn_cell.number_format = '@' # Text format
        
        ws.cell(row=current_row, column=4, value=r["qty"]).number_format = '#,##0'
        ws.cell(row=current_row, column=5, value=r["price"]).number_format = '$#,##0.00'
        ws.cell(row=current_row, column=6, value=r["taxable"]).number_format = '$#,##0.00'
        
        gst_cell = ws.cell(row=current_row, column=7, value=r["rate"])
        gst_cell.number_format = '0.0"%"'
        gst_cell.alignment = Alignment(horizontal="right")
        
        ws.cell(row=current_row, column=8, value=r["tax_amount"]).number_format = '$#,##0.00'
        ws.cell(row=current_row, column=9, value=r["total"]).number_format = '$#,##0.00'
        
        for c in range(1, 10):
            cell = ws.cell(row=current_row, column=c)
            cell.font = data_font
            cell.border = thin_border
        current_row += 1
        
    # Totals Row
    ws.cell(row=current_row, column=1, value="Total").font = bold_font
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left")
    
    ws.cell(row=current_row, column=6, value=total_taxable).number_format = '$#,##0.00'
    ws.cell(row=current_row, column=6).font = bold_font
    
    ws.cell(row=current_row, column=8, value=total_tax).number_format = '$#,##0.00'
    ws.cell(row=current_row, column=8).font = bold_font
    
    ws.cell(row=current_row, column=9, value=total_grand).number_format = '$#,##0.00'
    ws.cell(row=current_row, column=9).font = bold_font
    
    for c in range(1, 10):
        cell = ws.cell(row=current_row, column=c)
        cell.border = double_bottom_border
        cell.fill = summary_fill
        
    # Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row < 5:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 14
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_gst_pdf_report(month: str, data_rows: list, total_taxable: float, total_tax: float, total_grand: float) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor('#1F497D'),
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.HexColor('#595959'),
        spaceAfter=10
    )
    
    disclaimer_style = ParagraphStyle(
        'Disclaimer',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=8.5,
        textColor=colors.HexColor('#C00000'),
        leading=11,
        spaceAfter=15
    )
    
    table_text_style = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white
    )
    
    table_total_style = ParagraphStyle(
        'TableTotal',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10
    )

    elements = []
    elements.append(Paragraph("MediVision AI - GST Sales Report", title_style))
    elements.append(Paragraph(f"Report Period: {month}", subtitle_style))
    elements.append(Paragraph("This report is for reference only. Verify all figures and file directly through the official GST portal or your accountant.", disclaimer_style))
    
    table_data = [
        [
            Paragraph("Date", table_header_style),
            Paragraph("Medicine Name", table_header_style),
            Paragraph("HSN Code", table_header_style),
            Paragraph("Qty", table_header_style),
            Paragraph("Price", table_header_style),
            Paragraph("Taxable", table_header_style),
            Paragraph("GST Rate", table_header_style),
            Paragraph("GST Tax", table_header_style),
            Paragraph("Total Amount", table_header_style)
        ]
    ]
    
    for r in data_rows:
        table_data.append([
            Paragraph(r["date"], table_text_style),
            Paragraph(r["name"], table_text_style),
            Paragraph(r["hsn"] or "-", table_text_style),
            Paragraph(str(r["qty"]), table_text_style),
            Paragraph(f"${r['price']:.2f}", table_text_style),
            Paragraph(f"${r['taxable']:.2f}", table_text_style),
            Paragraph(f"{r['rate']:.1f}%", table_text_style),
            Paragraph(f"${r['tax_amount']:.2f}", table_text_style),
            Paragraph(f"${r['total']:.2f}", table_text_style)
        ])
        
    table_data.append([
        Paragraph("Total", table_total_style),
        Paragraph("", table_total_style),
        Paragraph("", table_total_style),
        Paragraph("", table_total_style),
        Paragraph("", table_total_style),
        Paragraph(f"${total_taxable:.2f}", table_total_style),
        Paragraph("", table_total_style),
        Paragraph(f"${total_tax:.2f}", table_total_style),
        Paragraph(f"${total_grand:.2f}", table_total_style)
    ])
    
    col_widths = [55, 135, 50, 25, 45, 55, 45, 55, 75]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F497D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
        ('TOPPADDING', (0, 0), (-1, 0), 4),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#D9D9D9')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E9EDF4')),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 4),
        ('TOPPADDING', (0, -1), (-1, -1), 4),
    ])
    
    for i in range(1, len(data_rows) + 1):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F2F2F2'))
            
    t.setStyle(t_style)
    elements.append(t)
    
    doc.build(elements)
    output.seek(0)
    return output.getvalue()


def generate_transactions_excel(inflow_rows: list, outflow_rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    
    # 1. Stock Inflow Sheet
    ws_in: Any = wb.active
    ws_in.title = "Stock Inflow"
    ws_in.views.sheetView[0].showGridLines = True
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    data_font = Font(name="Segoe UI", size=11)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    in_headers = ["Date", "Medicine Name", "Batch Number", "Quantity Added", "Purchase Price", "Added By (User/Role)"]
    for col_idx, text in enumerate(in_headers, 1):
        cell = ws_in.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for r_idx, r in enumerate(inflow_rows, 2):
        ws_in.cell(row=r_idx, column=1, value=r["date"]).alignment = Alignment(horizontal="center")
        ws_in.cell(row=r_idx, column=2, value=r["name"]).alignment = Alignment(horizontal="left")
        ws_in.cell(row=r_idx, column=3, value=r["batch"]).alignment = Alignment(horizontal="center")
        
        qty_cell = ws_in.cell(row=r_idx, column=4, value=r["qty"])
        qty_cell.number_format = '#,##0'
        qty_cell.alignment = Alignment(horizontal="right")
        
        price_cell = ws_in.cell(row=r_idx, column=5, value=r["price"])
        price_cell.number_format = '$#,##0.00'
        price_cell.alignment = Alignment(horizontal="right")
        
        ws_in.cell(row=r_idx, column=6, value=r["added_by"]).alignment = Alignment(horizontal="left")
        
        for col_idx in range(1, 7):
            cell = ws_in.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
    for col in ws_in.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws_in.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 2. Stock Outflow Sheet
    ws_out: Any = wb.create_sheet(title="Stock Outflow")
    ws_out.views.sheetView[0].showGridLines = True
    
    out_headers = ["Date", "Medicine Name", "Batch Number", "Quantity Sold", "Sale Price", "Total Amount", "Sold By (User/Role)"]
    out_fill = PatternFill(start_color="953734", end_color="953734", fill_type="solid")
    
    for col_idx, text in enumerate(out_headers, 1):
        cell = ws_out.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = out_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for r_idx, r in enumerate(outflow_rows, 2):
        ws_out.cell(row=r_idx, column=1, value=r["date"]).alignment = Alignment(horizontal="center")
        ws_out.cell(row=r_idx, column=2, value=r["name"]).alignment = Alignment(horizontal="left")
        ws_out.cell(row=r_idx, column=3, value=r["batch"]).alignment = Alignment(horizontal="center")
        
        qty_cell = ws_out.cell(row=r_idx, column=4, value=r["qty"])
        qty_cell.number_format = '#,##0'
        qty_cell.alignment = Alignment(horizontal="right")
        
        price_cell = ws_out.cell(row=r_idx, column=5, value=r["price"])
        price_cell.number_format = '$#,##0.00'
        price_cell.alignment = Alignment(horizontal="right")
        
        total_cell = ws_out.cell(row=r_idx, column=6, value=r["total"])
        total_cell.number_format = '$#,##0.00'
        total_cell.alignment = Alignment(horizontal="right")
        
        ws_out.cell(row=r_idx, column=7, value=r["sold_by"]).alignment = Alignment(horizontal="left")
        
        for col_idx in range(1, 8):
            cell = ws_out.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
    for col in ws_out.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws_out.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ── Schemas for Reports ──────────────────────────────────────────────────────

class MedicineGstConfig(BaseModel):
    hsn_code: str | None = None
    gst_rate: float = Field(..., ge=0.0, le=100.0, description="GST rate as percentage")

class GstReportRequest(BaseModel):
    month: str = Field(..., description="Month in YYYY-MM format")
    medicines_config: dict[str, MedicineGstConfig] = Field(..., description="Map of medicine_id (str) to config")
    format: str = Field(default="pdf", description="Format: pdf or excel")


# ── Endpoints for Reports ────────────────────────────────────────────────────

@router.get(
    "/gst-report/medicines/{month}",
    summary="Get unique medicines sold in a given month",
)
def get_gst_report_medicines(
    month: str,
    db: Session = Depends(get_db),
    current_user: SupabaseUser = Depends(get_current_user),
):
    if current_user.role not in ("admin", "pharmacist"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Staff cannot access Monthly Business Overview."
        )
    if not re.match(r"^\d{4}-\d{2}$", month):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Month must be in YYYY-MM format (e.g., 2026-07)"
        )
    
    year, month_num = map(int, month.split("-"))
    import calendar
    last_day = calendar.monthrange(year, month_num)[1]
    start_dt = datetime(year, month_num, 1, 0, 0, 0)
    end_dt = datetime(year, month_num, last_day, 23, 59, 59, 999999)

    results = (
        db.query(Medicine.id, Medicine.name)
        .select_from(SaleItem)
        .join(Sale, SaleItem.sale_id == Sale.id)
        .join(Medicine, SaleItem.medicine_id == Medicine.id)
        .filter(Sale.sold_at >= start_dt, Sale.sold_at <= end_dt)
        .distinct()
        .all()
    )
    return [{"id": r[0], "name": r[1]} for r in results]


@router.post(
    "/gst-report",
    summary="Generate GST sales report for a month",
)
def generate_gst_report(
    request_data: GstReportRequest,
    db: Session = Depends(get_db),
    current_user: SupabaseUser = Depends(get_current_user),
):
    if current_user.role not in ("admin", "pharmacist"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Staff cannot access Monthly Business Overview."
        )
    month = request_data.month
    if not re.match(r"^\d{4}-\d{2}$", month):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Month must be in YYYY-MM format (e.g., 2026-07)"
        )
    
    year, month_num = map(int, month.split("-"))
    import calendar
    last_day = calendar.monthrange(year, month_num)[1]
    start_dt = datetime(year, month_num, 1, 0, 0, 0)
    end_dt = datetime(year, month_num, last_day, 23, 59, 59, 999999)

    items = (
        db.query(SaleItem)
        .join(Sale, SaleItem.sale_id == Sale.id)
        .filter(Sale.sold_at >= start_dt, Sale.sold_at <= end_dt)
        .order_by(Sale.sold_at.asc())
        .all()
    )

    data_rows = []
    total_taxable = 0.0
    total_tax = 0.0
    total_grand = 0.0

    for item in items:
        med_name = item.medicine.name if item.medicine else "[Deleted Medicine]"
        med_id_str = str(item.medicine_id) if item.medicine_id else ""
        config = request_data.medicines_config.get(med_id_str)
        
        hsn = config.hsn_code if config else ""
        gst_rate = config.gst_rate if config else 0.0
        
        taxable = item.quantity_sold * float(item.sale_price)
        tax_amount = taxable * (gst_rate / 100.0)
        total = taxable + tax_amount
        
        total_taxable += taxable
        total_tax += tax_amount
        total_grand += total
        
        data_rows.append({
            "date": item.sale.sold_at.strftime("%Y-%m-%d"),
            "name": med_name,
            "hsn": hsn or "",
            "qty": item.quantity_sold,
            "price": float(item.sale_price),
            "taxable": taxable,
            "rate": gst_rate,
            "tax_amount": tax_amount,
            "total": total
        })

    if request_data.format == "excel":
        file_bytes = generate_gst_excel_report(month, data_rows, total_taxable, total_tax, total_grand)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"GST_Report_{month}.xlsx"
    else:
        file_bytes = generate_gst_pdf_report(month, data_rows, total_taxable, total_tax, total_grand)
        media_type = "application/pdf"
        filename = f"GST_Report_{month}.pdf"

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


@router.get(
    "/transactions/export",
    summary="Export all inflow and outflow transactions as Excel workbook",
)
def export_transactions(
    start_date: str | None = None,
    end_date: str | None = None,
    month: str | None = None,
    db: Session = Depends(get_db),
    current_user: SupabaseUser = Depends(get_current_user),
):
    if current_user.role not in ("admin", "pharmacist"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Staff cannot access Monthly Business Overview."
        )
        
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="start_date must be in YYYY-MM-DD format")
    else:
        start_dt = None

    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, microsecond=999999)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="end_date must be in YYYY-MM-DD format")
    else:
        end_dt = None

    if not start_dt and not end_dt:
        target_month = month or datetime.now().strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", target_month):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="month must be in YYYY-MM format")
        year, month_num = map(int, target_month.split("-"))
        import calendar
        last_day = calendar.monthrange(year, month_num)[1]
        start_dt = datetime(year, month_num, 1, 0, 0, 0)
        end_dt = datetime(year, month_num, last_day, 23, 59, 59, 999999)

    # Inflow
    inflow_records = (
        db.query(ExtractionRecord)
        .filter(
            ExtractionRecord.status == "done",
            ExtractionRecord.confirmed_at >= start_dt,
            ExtractionRecord.confirmed_at <= end_dt
        )
        .order_by(ExtractionRecord.confirmed_at.asc())
        .all()
    )

    inflow_rows = []
    for rec in inflow_records:
        name = ""
        batch = ""
        price = 0.0
        if rec.final_values:
            try:
                fv = json.loads(rec.final_values)
                name = fv.get("medicine_name") or ""
                batch = fv.get("batch_number") or ""
                price = float(fv.get("purchase_price") or 0.0)
            except Exception:
                pass

        if not name and rec.medicine:
            name = rec.medicine.name
            batch = rec.medicine.batch_number
            price = float(rec.medicine.purchase_price)

        qty = 0
        if rec.medicine_id and rec.confirmed_at:
            audit_log = db.query(AuditLog).filter(
                AuditLog.medicine_id == rec.medicine_id,
                AuditLog.action.in_(["created", "quantity_updated"]),
                AuditLog.timestamp >= rec.confirmed_at - timedelta(seconds=15),
                AuditLog.timestamp <= rec.confirmed_at + timedelta(seconds=15)
            ).first()

            if audit_log:
                if audit_log.action == "created":
                    try:
                        val = json.loads(audit_log.new_value)
                        qty = val.get("quantity") or 0
                    except Exception:
                        pass
                elif audit_log.action == "quantity_updated":
                    try:
                        qty = int(audit_log.new_value) - int(audit_log.old_value)
                    except Exception:
                        pass

        inflow_rows.append({
            "date": rec.confirmed_at.strftime("%Y-%m-%d") if rec.confirmed_at else "",
            "name": name,
            "batch": batch,
            "qty": qty,
            "price": price,
            "added_by": rec.confirmed_by or ""
        })

    # Outflow
    outflow_items = (
        db.query(SaleItem)
        .join(Sale, SaleItem.sale_id == Sale.id)
        .filter(Sale.sold_at >= start_dt, Sale.sold_at <= end_dt)
        .order_by(Sale.sold_at.asc())
        .all()
    )

    outflow_rows = []
    for item in outflow_items:
        med_name = item.medicine.name if item.medicine else "[Deleted Medicine]"
        batch = item.medicine.batch_number if item.medicine else "-"
        outflow_rows.append({
            "date": item.sale.sold_at.strftime("%Y-%m-%d"),
            "name": med_name,
            "batch": batch,
            "qty": item.quantity_sold,
            "price": float(item.sale_price),
            "total": float(item.line_total),
            "sold_by": item.sale.sold_by or ""
        })

    file_bytes = generate_transactions_excel(inflow_rows, outflow_rows)
    assert start_dt is not None
    assert end_dt is not None
    filename = f"Transactions_Export_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


# ── Additional PDF / Excel Helpers for Sales and Expiry Reports ────────────────

def generate_sales_pdf_report(
    month: str,
    rent: float,
    bills: float,
    salaries: float,
    other: float,
    other_rev: float,
    computed_sales: float,
    total_revenue: float,
    total_costs: float,
    inventory_investment: float,
    net_profit: float,
    roi: float,
    estimated_margin: float,
) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor('#0A74DA'),
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.HexColor('#595959'),
        spaceAfter=10
    )
    
    disclaimer_style = ParagraphStyle(
        'Disclaimer',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=8.5,
        textColor=colors.HexColor('#E64A19'),
        leading=11,
        spaceAfter=15
    )
    
    label_style = ParagraphStyle(
        'Label',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.HexColor('#333333')
    )
    
    val_style = ParagraphStyle(
        'Val',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#111111')
    )
    
    kpi_title_style = ParagraphStyle(
        'KpiTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    kpi_val_style = ParagraphStyle(
        'KpiVal',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.white,
        alignment=1 # Center
    )

    elements = []
    elements.append(Paragraph("MediVision AI - Monthly Sales P&L Ledger", title_style))
    elements.append(Paragraph(f"Report Period: {month}", subtitle_style))
    elements.append(Paragraph("This report is for pharmacy operational review. Verify all transaction records with physical inventories and direct billing registries.", disclaimer_style))
    
    kpi_data = [
        [
            Paragraph("Total Revenue", kpi_title_style),
            Paragraph("Total Costs", kpi_title_style),
            Paragraph("Net Profit/Loss", kpi_title_style),
            Paragraph("Return on Investment", kpi_title_style)
        ],
        [
            Paragraph(f"INR {total_revenue:,.2f}", kpi_val_style),
            Paragraph(f"INR {total_costs:,.2f}", kpi_val_style),
            Paragraph(f"INR {net_profit:,.2f}", kpi_val_style),
            Paragraph(f"{roi:.1f}%", kpi_val_style)
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[130, 130, 130, 130])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#0A74DA')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#06b6d4')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#0A74DA')),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 20))
    
    pl_data = [
        [Paragraph("<b>Category / Line Item</b>", label_style), Paragraph("<b>Value</b>", label_style)],
        [Paragraph("Computed Sales Revenue", label_style), Paragraph(f"INR {computed_sales:,.2f}", val_style)],
        [Paragraph("Other Overrides / Custom Income", label_style), Paragraph(f"INR {other_rev:,.2f}", val_style)],
        [Paragraph("<b>Gross Monthly Revenue</b>", label_style), Paragraph(f"<b>INR {total_revenue:,.2f}</b>", val_style)],
        [Paragraph("", label_style), Paragraph("", val_style)],
        [Paragraph("Fixed Rent Cost", label_style), Paragraph(f"INR {rent:,.2f}", val_style)],
        [Paragraph("Electricity & Utility Bills", label_style), Paragraph(f"INR {bills:,.2f}", val_style)],
        [Paragraph("Staff Salaries", label_style), Paragraph(f"INR {salaries:,.2f}", val_style)],
        [Paragraph("Other Fixed Operational Overheads", label_style), Paragraph(f"INR {other:,.2f}", val_style)],
        [Paragraph("<b>Total Monthly Costs</b>", label_style), Paragraph(f"<b>INR {total_costs:,.2f}</b>", val_style)],
        [Paragraph("", label_style), Paragraph("", val_style)],
        [Paragraph("Current Inventory Investment", label_style), Paragraph(f"INR {inventory_investment:,.2f}", val_style)],
        [Paragraph("Estimated Sales Margin", label_style), Paragraph(f"INR {estimated_margin:,.2f}", val_style)],
        [Paragraph("<b>Net Monthly P&L</b>", label_style), Paragraph(f"<b>INR {net_profit:,.2f}</b>", val_style)]
    ]
    
    pl_table = Table(pl_data, colWidths=[360, 160])
    pl_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (0, 3), (1, 3), colors.HexColor('#F1F5F9')),
        ('BACKGROUND', (0, 9), (1, 9), colors.HexColor('#F1F5F9')),
        ('BACKGROUND', (0, 13), (1, 13), colors.HexColor('#ECFDF5') if net_profit >= 0 else colors.HexColor('#FEF2F2')),
    ]))
    elements.append(pl_table)
    
    doc.build(elements)
    return output.getvalue()


def generate_sales_excel_report(
    month: str,
    rent: float,
    bills: float,
    salaries: float,
    other: float,
    other_rev: float,
    computed_sales: float,
    total_revenue: float,
    total_costs: float,
    inventory_investment: float,
    net_profit: float,
    roi: float,
    estimated_margin: float,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws: Any = wb.active
    ws.title = "P&L Summary"
    ws.views.sheetView[0].showGridLines = True
    
    title_font = Font(name="Segoe UI", size=14, bold=True, color="0A74DA")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    label_font = Font(name="Segoe UI", size=10, bold=True)
    data_font = Font(name="Segoe UI", size=10)
    disclaimer_font = Font(name="Segoe UI", size=9, italic=True, color="C00000")
    
    header_fill = PatternFill(start_color="0A74DA", end_color="0A74DA", fill_type="solid")
    summary_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    ws.cell(row=1, column=1, value="MediVision AI - Monthly Sales P&L Ledger").font = title_font
    ws.cell(row=2, column=1, value=f"Report Period: {month}").font = label_font
    ws.cell(row=3, column=1, value="Operational Monthly P&L Ledger analysis.").font = disclaimer_font
    
    ws.cell(row=5, column=1, value="Category / Line Item").font = header_font
    ws.cell(row=5, column=1).fill = header_fill
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=5, column=2, value="Value").font = header_font
    ws.cell(row=5, column=2).fill = header_fill
    ws.cell(row=5, column=2).alignment = Alignment(horizontal="right")
    
    rows = [
        ("Computed Sales Revenue", computed_sales),
        ("Other Overrides / Custom Income", other_rev),
        ("Gross Monthly Revenue", total_revenue),
        ("", None),
        ("Fixed Rent Cost", rent),
        ("Electricity & Utility Bills", bills),
        ("Staff Salaries", salaries),
        ("Other Fixed Operational Overheads", other),
        ("Total Monthly Costs", total_costs),
        ("", None),
        ("Current Inventory Investment", inventory_investment),
        ("Estimated Sales Margin", estimated_margin),
        ("Return on Investment", f"{roi:.1f}%"),
        ("Net Monthly P&L", net_profit),
    ]
    
    curr = 6
    for name, val in rows:
        ws.cell(row=curr, column=1, value=name).font = Font(name="Segoe UI", size=10, bold=True) if "Total" in name or "Gross" in name or "Net" in name else data_font
        ws.cell(row=curr, column=1).border = thin_border
        
        val_cell = ws.cell(row=curr, column=2, value=val)
        val_cell.border = thin_border
        if val is not None:
            if isinstance(val, (int, float)):
                val_cell.number_format = '[$₹-3601] #,##0.00'
                val_cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, str) and "%" in val:
                val_cell.alignment = Alignment(horizontal="right")
            val_cell.font = Font(name="Segoe UI", size=10, bold=True) if "Total" in name or "Gross" in name or "Net" in name else data_font
            
        if "Gross" in name or "Total" in name:
            ws.cell(row=curr, column=1).fill = summary_fill
            ws.cell(row=curr, column=2).fill = summary_fill
        elif "Net" in name:
            fill = green_fill if net_profit >= 0 else red_fill
            ws.cell(row=curr, column=1).fill = fill
            ws.cell(row=curr, column=2).fill = fill
            
        curr += 1
        
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_expiry_pdf_report(today, expired_items: list, expiring_items: list) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor('#E64A19'),
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.HexColor('#595959'),
        spaceAfter=10
    )
    
    disclaimer_style = ParagraphStyle(
        'Disclaimer',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=8.5,
        textColor=colors.HexColor('#D32F2F'),
        leading=11,
        spaceAfter=15
    )
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=colors.HexColor('#1E293B'),
        spaceBefore=12,
        spaceAfter=6
    )
    
    table_text_style = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white
    )
    
    table_total_style = ParagraphStyle(
        'TableTotal',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10
    )

    elements = []
    elements.append(Paragraph("MediVision AI - Expiry Write-off &amp; Inventory Risk Report", title_style))
    elements.append(Paragraph(f"Report Generated: {today.strftime('%Y-%m-%d')}", subtitle_style))
    elements.append(Paragraph("CRITICAL SAFETY AND FINANCE CONTROL: Items identified below as 'Expired' must be immediately quarantined and removed from active stock. Expiring items represent financial risk and should be prioritized for distribution or discount campaigns.", disclaimer_style))
    
    total_expired_val = sum(x.quantity * x.purchase_price for x in expired_items)
    total_expiring_val = sum(x.quantity * x.purchase_price for x in expiring_items)
    
    summary_data = [
        [
            Paragraph("Category", table_header_style),
            Paragraph("SKUs Count", table_header_style),
            Paragraph("Total Quantity", table_header_style),
            Paragraph("Estimated Value at Risk", table_header_style)
        ],
        [
            Paragraph("<b>Expired Stock</b> (Write-off)", table_text_style),
            Paragraph(str(len(expired_items)), table_text_style),
            Paragraph(str(sum(x.quantity for x in expired_items)), table_text_style),
            Paragraph(f"INR {total_expired_val:,.2f}", table_text_style)
        ],
        [
            Paragraph("<b>Expiring soon</b> (Next 90 Days)", table_text_style),
            Paragraph(str(len(expiring_items)), table_text_style),
            Paragraph(str(sum(x.quantity for x in expiring_items)), table_text_style),
            Paragraph(f"INR {total_expiring_val:,.2f}", table_text_style)
        ]
    ]
    summary_table = Table(summary_data, colWidths=[160, 100, 100, 160])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (3, 1), colors.HexColor('#FEF2F2')),
        ('BACKGROUND', (0, 2), (3, 2), colors.HexColor('#FFFBEB')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 15))
    
    elements.append(Paragraph("Quarantine &amp; Write-off Registry (Expired)", section_title_style))
    if expired_items:
        expired_data = [
            [
                Paragraph("Medicine Name", table_header_style),
                Paragraph("Batch", table_header_style),
                Paragraph("Expiry Date", table_header_style),
                Paragraph("Qty", table_header_style),
                Paragraph("Purchase Price", table_header_style),
                Paragraph("Write-off Value", table_header_style),
                Paragraph("Location", table_header_style)
            ]
        ]
        for item in expired_items:
            item_val = item.quantity * item.purchase_price
            expired_data.append([
                Paragraph(item.name, table_text_style),
                Paragraph(item.batch_number or "-", table_text_style),
                Paragraph(item.expiry_date.strftime("%Y-%m-%d") if item.expiry_date else "-", table_text_style),
                Paragraph(str(item.quantity), table_text_style),
                Paragraph(f"INR {item.purchase_price:.2f}", table_text_style),
                Paragraph(f"INR {item_val:.2f}", table_text_style),
                Paragraph(item.storage_location or "-", table_text_style)
            ])
        
        expired_data.append([
            Paragraph("Total Expired Loss", table_total_style),
            Paragraph("", table_total_style),
            Paragraph("", table_total_style),
            Paragraph("", table_total_style),
            Paragraph("", table_total_style),
            Paragraph(f"INR {total_expired_val:.2f}", table_total_style),
            Paragraph("", table_total_style)
        ])
        
        t_exp = Table(expired_data, colWidths=[120, 50, 60, 30, 70, 75, 115])
        t_exp.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F8FAFC')),
        ]))
        elements.append(t_exp)
    else:
        elements.append(Paragraph("No expired stock items registered in inventory.", table_text_style))
        
    elements.append(Spacer(1, 15))
    
    elements.append(Paragraph("Near Expiry Risk Warning Registry (Next 90 Days)", section_title_style))
    if expiring_items:
        expiring_data = [
            [
                Paragraph("Medicine Name", table_header_style),
                Paragraph("Batch", table_header_style),
                Paragraph("Expiry Date", table_header_style),
                Paragraph("Qty", table_header_style),
                Paragraph("Purchase Price", table_header_style),
                Paragraph("Value at Risk", table_header_style),
                Paragraph("Location", table_header_style)
            ]
        ]
        for item in expiring_items:
            item_val = item.quantity * item.purchase_price
            expiring_data.append([
                Paragraph(item.name, table_text_style),
                Paragraph(item.batch_number or "-", table_text_style),
                Paragraph(item.expiry_date.strftime("%Y-%m-%d") if item.expiry_date else "-", table_text_style),
                Paragraph(str(item.quantity), table_text_style),
                Paragraph(f"INR {item.purchase_price:.2f}", table_text_style),
                Paragraph(f"INR {item_val:.2f}", table_text_style),
                Paragraph(item.storage_location or "-", table_text_style)
            ])
            
        expiring_data.append([
            Paragraph("Total Expiring Risk", table_total_style),
            Paragraph("", table_total_style),
            Paragraph("", table_total_style),
            Paragraph("", table_total_style),
            Paragraph("", table_total_style),
            Paragraph(f"INR {total_expiring_val:.2f}", table_total_style),
            Paragraph("", table_total_style)
        ])
        
        t_nexp = Table(expiring_data, colWidths=[120, 50, 60, 30, 70, 75, 115])
        t_nexp.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D97706')),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F8FAFC')),
        ]))
        elements.append(t_nexp)
    else:
        elements.append(Paragraph("No items expiring in the next 90 days.", table_text_style))
        
    doc.build(elements)
    return output.getvalue()


def generate_expiry_excel_report(today, expired_items: list, expiring_items: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws_sum: Any = wb.active
    ws_sum.title = "Overview"
    ws_sum.views.sheetView[0].showGridLines = True
    
    title_font = Font(name="Segoe UI", size=14, bold=True, color="E64A19")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    label_font = Font(name="Segoe UI", size=11, bold=True)
    data_font = Font(name="Segoe UI", size=11)
    disclaimer_font = Font(name="Segoe UI", size=9, italic=True, color="C00000")
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    expired_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    expiring_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    ws_sum.cell(row=1, column=1, value="MediVision AI - Expiry Audit Summary").font = title_font
    ws_sum.cell(row=2, column=1, value=f"Report Generated: {today.strftime('%Y-%m-%d')}").font = label_font
    ws_sum.cell(row=3, column=1, value="Critical Safety Control: Quarantined & soon-expiring inventories.").font = disclaimer_font
    
    headers = ["Category", "SKUs Count", "Total Quantity", "Estimated Value at Risk"]
    for idx, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=5, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    total_expired_val = sum(x.quantity * x.purchase_price for x in expired_items)
    total_expiring_val = sum(x.quantity * x.purchase_price for x in expiring_items)
    
    ws_sum.cell(row=6, column=1, value="Expired Stock (Write-off)").font = label_font
    ws_sum.cell(row=6, column=2, value=len(expired_items)).number_format = '#,##0'
    ws_sum.cell(row=6, column=3, value=sum(x.quantity for x in expired_items)).number_format = '#,##0'
    ws_sum.cell(row=6, column=4, value=total_expired_val).number_format = '[$₹-3601] #,##0.00'
    
    ws_sum.cell(row=7, column=1, value="Expiring Stock (Next 90 Days)").font = label_font
    ws_sum.cell(row=7, column=2, value=len(expiring_items)).number_format = '#,##0'
    ws_sum.cell(row=7, column=3, value=sum(x.quantity for x in expiring_items)).number_format = '#,##0'
    ws_sum.cell(row=7, column=4, value=total_expiring_val).number_format = '[$₹-3601] #,##0.00'
    
    for r in (6, 7):
        for c in range(1, 5):
            cell = ws_sum.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = data_font
            cell.fill = expired_fill if r == 6 else expiring_fill
            if c > 1:
                cell.alignment = Alignment(horizontal="right")
                
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 15
    ws_sum.column_dimensions["D"].width = 25
    
    ws_exp: Any = wb.create_sheet(title="Expired Quarantine List")
    ws_exp.views.sheetView[0].showGridLines = True
    
    det_headers = ["Medicine Name", "Manufacturer", "Batch", "Expiry Date", "Quantity", "Purchase Price", "Write-off Cost", "Location"]
    red_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    
    for idx, h in enumerate(det_headers, 1):
        cell = ws_exp.cell(row=1, column=idx, value=h)
        cell.font = header_font
        cell.fill = red_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    curr_row = 2
    for item in expired_items:
        ws_exp.cell(row=curr_row, column=1, value=item.name).border = thin_border
        ws_exp.cell(row=curr_row, column=2, value=item.manufacturer or "").border = thin_border
        ws_exp.cell(row=curr_row, column=3, value=item.batch_number or "").border = thin_border
        ws_exp.cell(row=curr_row, column=4, value=item.expiry_date.strftime("%Y-%m-%d") if item.expiry_date else "").border = thin_border
        ws_exp.cell(row=curr_row, column=4).alignment = Alignment(horizontal="center")
        
        ws_exp.cell(row=curr_row, column=5, value=item.quantity).number_format = '#,##0'
        ws_exp.cell(row=curr_row, column=5).alignment = Alignment(horizontal="right")
        
        ws_exp.cell(row=curr_row, column=6, value=item.purchase_price).number_format = '[$₹-3601] #,##0.00'
        ws_exp.cell(row=curr_row, column=6).alignment = Alignment(horizontal="right")
        
        ws_exp.cell(row=curr_row, column=7, value=item.quantity * item.purchase_price).number_format = '[$₹-3601] #,##0.00'
        ws_exp.cell(row=curr_row, column=7).alignment = Alignment(horizontal="right")
        
        ws_exp.cell(row=curr_row, column=8, value=item.storage_location or "").border = thin_border
        
        for c in range(1, 9):
            ws_exp.cell(row=curr_row, column=c).border = thin_border
            ws_exp.cell(row=curr_row, column=c).font = data_font
            
        curr_row += 1
        
    for col in ws_exp.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_exp.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
        
    ws_nexp: Any = wb.create_sheet(title="Expiring Soon Warning List")
    ws_nexp.views.sheetView[0].showGridLines = True
    
    amber_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    for idx, h in enumerate(det_headers, 1):
        cell = ws_nexp.cell(row=1, column=idx, value=h)
        cell.font = header_font
        cell.fill = amber_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    curr_row = 2
    for item in expiring_items:
        ws_nexp.cell(row=curr_row, column=1, value=item.name).border = thin_border
        ws_nexp.cell(row=curr_row, column=2, value=item.manufacturer or "").border = thin_border
        ws_nexp.cell(row=curr_row, column=3, value=item.batch_number or "").border = thin_border
        ws_nexp.cell(row=curr_row, column=4, value=item.expiry_date.strftime("%Y-%m-%d") if item.expiry_date else "").border = thin_border
        ws_nexp.cell(row=curr_row, column=4).alignment = Alignment(horizontal="center")
        
        ws_nexp.cell(row=curr_row, column=5, value=item.quantity).number_format = '#,##0'
        ws_nexp.cell(row=curr_row, column=5).alignment = Alignment(horizontal="right")
        
        ws_nexp.cell(row=curr_row, column=6, value=item.purchase_price).number_format = '[$₹-3601] #,##0.00'
        ws_nexp.cell(row=curr_row, column=6).alignment = Alignment(horizontal="right")
        
        ws_nexp.cell(row=curr_row, column=7, value=item.quantity * item.purchase_price).number_format = '[$₹-3601] #,##0.00'
        ws_nexp.cell(row=curr_row, column=7).alignment = Alignment(horizontal="right")
        
        ws_nexp.cell(row=curr_row, column=8, value=item.storage_location or "").border = thin_border
        
        for c in range(1, 9):
            ws_nexp.cell(row=curr_row, column=c).border = thin_border
            ws_nexp.cell(row=curr_row, column=c).font = data_font
            
        curr_row += 1
        
    for col in ws_nexp.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_nexp.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ── Report Generation Endpoints ──────────────────────────────────────────────

@router.get(
    "/gst-report/summary/{month}",
    summary="Get GST tax totals and summaries for a month",
)
def get_gst_report_summary(
    month: str,
    db: Session = Depends(get_db),
    current_user: SupabaseUser = Depends(get_current_user),
):
    if current_user.role not in ("admin", "pharmacist"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Staff cannot access report summaries."
        )
    if not re.match(r"^\d{4}-\d{2}$", month):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Month must be in YYYY-MM format (e.g., 2026-07)"
        )
    
    year, month_num = map(int, month.split("-"))
    import calendar
    last_day = calendar.monthrange(year, month_num)[1]
    start_dt = datetime(year, month_num, 1, 0, 0, 0)
    end_dt = datetime(year, month_num, last_day, 23, 59, 59, 999999)

    items = (
        db.query(SaleItem)
        .join(Sale, SaleItem.sale_id == Sale.id)
        .filter(Sale.sold_at >= start_dt, Sale.sold_at <= end_dt)
        .all()
    )

    unique_medicines = set()
    total_taxable = 0.0
    total_tax = 0.0
    total_grand = 0.0

    for item in items:
        unique_medicines.add(item.medicine_id)
        taxable = item.quantity_sold * float(item.sale_price)
        gst_rate = 18.0  # Standard GST assumption for preview summaries
        tax_amount = taxable * (gst_rate / 100.0)
        total = taxable + tax_amount

        total_taxable += taxable
        total_tax += tax_amount
        total_grand += total

    return {
        "unique_medicines_count": len(unique_medicines),
        "total_taxable_value": total_taxable,
        "estimated_gst_tax": total_tax,
        "total_grand_amount": total_grand
    }


@router.get(
    "/sales-report",
    summary="Download monthly revenue P&L ledger report",
)
def download_sales_report(
    month: str,
    format: str = "pdf",
    db: Session = Depends(get_db),
    current_user: SupabaseUser = Depends(get_current_user),
):
    if current_user.role not in ("admin", "pharmacist"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Staff cannot access report generation."
        )
    if not re.match(r"^\d{4}-\d{2}$", month):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Month must be in YYYY-MM format (e.g., 2026-07)"
        )

    record = (
        db.query(MonthlyFinance)
        .filter(MonthlyFinance.month == month)
        .first()
    )

    rent = float(record.rent) if record else 0.0
    bills = float(record.electricity_and_bills) if record else 0.0
    salaries = float(record.staff_salaries) if record else 0.0
    other = float(record.other_expenses) if record else 0.0
    other_rev = float(record.other_revenue) if record else 0.0

    computed_sales = get_computed_sales_for_month(db, month)
    total_revenue = computed_sales + other_rev
    total_costs = rent + bills + salaries + other

    inventory_val = db.query(func.sum(Medicine.quantity * Medicine.purchase_price)).filter(Medicine.quantity > 0).scalar()
    current_inventory_investment = float(inventory_val) if inventory_val is not None else 0.0

    net_profit = total_revenue - total_costs
    roi = (net_profit / current_inventory_investment) * 100.0 if current_inventory_investment > 0.0 else 0.0

    # Estimated Margin
    year_str, month_str = month.split("-")
    year = int(year_str)
    month_num = int(month_str)
    start_dt = datetime(year, month_num, 1, 0, 0, 0)
    import calendar
    last_day = calendar.monthrange(year, month_num)[1]
    end_dt = datetime(year, month_num, last_day, 23, 59, 59, 999999)

    margin_val = (
        db.query(func.sum((SaleItem.sale_price - Medicine.purchase_price) * SaleItem.quantity_sold))
        .select_from(SaleItem)
        .join(Sale, SaleItem.sale_id == Sale.id)
        .join(Medicine, SaleItem.medicine_id == Medicine.id)
        .filter(Sale.sold_at >= start_dt, Sale.sold_at <= end_dt)
        .scalar()
    )
    estimated_margin = float(margin_val) if margin_val is not None else 0.0

    if format.lower() == "excel":
        file_bytes = generate_sales_excel_report(month, rent, bills, salaries, other, other_rev, computed_sales, total_revenue, total_costs, current_inventory_investment, net_profit, roi, estimated_margin)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"Sales_PL_Report_{month}.xlsx"
    else:
        file_bytes = generate_sales_pdf_report(month, rent, bills, salaries, other, other_rev, computed_sales, total_revenue, total_costs, current_inventory_investment, net_profit, roi, estimated_margin)
        media_type = "application/pdf"
        filename = f"Sales_PL_Report_{month}.pdf"

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


@router.get(
    "/expiry-report",
    summary="Download expiry write-off audit report",
)
def download_expiry_report(
    format: str = "pdf",
    db: Session = Depends(get_db),
    current_user: SupabaseUser = Depends(get_current_user),
):
    if current_user.role not in ("admin", "pharmacist"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Staff cannot access report generation."
        )

    today = date.today()
    ninety_days_later = today + timedelta(days=90)

    expired_items = (
        db.query(Medicine)
        .filter(
            Medicine.expiry_date < today,
            Medicine.quantity > 0
        )
        .order_by(Medicine.expiry_date.asc())
        .all()
    )

    expiring_items = (
        db.query(Medicine)
        .filter(
            Medicine.expiry_date >= today,
            Medicine.expiry_date <= ninety_days_later,
            Medicine.quantity > 0
        )
        .order_by(Medicine.expiry_date.asc())
        .all()
    )

    if format.lower() == "excel":
        file_bytes = generate_expiry_excel_report(today, expired_items, expiring_items)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"Expiry_Audit_Report_{today.strftime('%Y-%m-%d')}.xlsx"
    else:
        file_bytes = generate_expiry_pdf_report(today, expired_items, expiring_items)
        media_type = "application/pdf"
        filename = f"Expiry_Audit_Report_{today.strftime('%Y-%m-%d')}.pdf"

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


